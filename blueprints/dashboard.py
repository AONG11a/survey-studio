"""Dashboard blueprint — detailed per-form analytics.

Endpoints:
  /forms/<id>/dashboard            GET   dashboard HTML page
  /api/forms/<id>/analytics        GET   JSON analytics (charts + summary)
  /api/forms/<id>/responses        GET   JSON list of individual responses
  /api/forms/<id>/responses/<rid>  GET   JSON single response detail
  /api/forms/<id>/export.csv       GET   CSV export of all responses
  /api/forms/<id>/export.xlsx      GET   Excel export of all responses
"""
import csv
import io
import json
from collections import Counter
from datetime import datetime, timedelta
from flask import (
    Blueprint, render_template, request, jsonify, Response as FlaskResponse,
)
from flask_login import login_required, current_user
from models import db, Form, Question, Response, Answer
from authz import require_form_ownership

dashboard_bp = Blueprint("dashboard", __name__)


def _word_freq(texts, top_n=20):
    """Thai-friendly word frequency. Splits on whitespace + common punctuation."""
    import re
    counter = Counter()
    stop = {
        # minimal Thai stop set
        "และ", "หรือ", "ของ", "ที่", "ใน", "การ", "เป็น", "ได้", "มี", "ไม่",
        "ก็", "ให้", "และก็", "จะ", "ใช่", "the", "a", "an", "of", "to",
        "and", "or", "is", "are", "was", "were", "i", "you", "we", "they",
    }
    for t in texts:
        if not t:
            continue
        words = re.findall(r"\w+", t.lower(), flags=re.UNICODE)
        for w in words:
            if len(w) < 2 or w in stop:
                continue
            counter[w] += 1
    return counter.most_common(top_n)


def _summarize_question(q, answers):
    """Build a type-appropriate summary dict for one question."""
    summary = {
        "question_id": q.id,
        "text": q.text,
        "type": q.type,
        "answer_count": len(answers),
    }

    if q.type in ("multiple", "dropdown"):
        options = json.loads(q.options_json) if q.options_json else []
        counts = Counter()
        for a in answers:
            v = a.parsed_value()
            if v:
                counts[str(v)] += 1
        total = sum(counts.values()) or 1
        summary["options"] = options
        summary["counts"] = {o: counts.get(o, 0) for o in options}
        summary["percent"] = {o: round(counts.get(o, 0) / total * 100, 1) for o in options}
        summary["chart"] = "pie"

    elif q.type == "checkbox":
        options = json.loads(q.options_json) if q.options_json else []
        counts = Counter()
        for a in answers:
            v = a.parsed_value()
            if isinstance(v, list):
                for item in v:
                    counts[str(item)] += 1
        total = len(answers) or 1
        summary["options"] = options
        summary["counts"] = {o: counts.get(o, 0) for o in options}
        summary["percent"] = {o: round(counts.get(o, 0) / total * 100, 1) for o in options}
        summary["chart"] = "bar"
        summary["note"] = "เปอร์เซ็นต์อาจเกิน 100% เพราะเลือกได้หลายข้อ"

    elif q.type == "scale":
        nums = []
        for a in answers:
            v = a.parsed_value()
            try:
                nums.append(int(v))
            except (TypeError, ValueError):
                pass
        summary["min"] = q.scale_min
        summary["max"] = q.scale_max
        summary["count"] = len(nums)
        if nums:
            s_nums = sorted(nums)
            n = len(s_nums)
            summary["mean"] = round(sum(nums) / n, 2)
            if n % 2 == 1:
                summary["median"] = s_nums[n // 2]
            else:
                summary["median"] = round((s_nums[n // 2 - 1] + s_nums[n // 2]) / 2, 2)
            summary["distribution"] = {
                str(i): nums.count(i) for i in range(q.scale_min, q.scale_max + 1)
            }
        else:
            summary["mean"] = None
            summary["median"] = None
            summary["distribution"] = {}
        summary["chart"] = "histogram"

    elif q.type in ("short", "paragraph"):
        texts = [a.parsed_value() for a in answers if a.parsed_value()]
        summary["responses"] = texts
        summary["word_freq"] = _word_freq(texts)
        summary["chart"] = "text"

    elif q.type == "date":
        dates = []
        for a in answers:
            v = a.parsed_value()
            try:
                dates.append(str(v)[:10])
            except (TypeError, ValueError):
                pass
        counts = Counter(dates)
        summary["distribution"] = dict(sorted(counts.items()))
        summary["chart"] = "bar"

    elif q.type == "time":
        times = []
        for a in answers:
            v = a.parsed_value()
            try:
                times.append(str(v)[:5])
            except (TypeError, ValueError):
                pass
        hours = Counter(t.split(":")[0] for t in times if t)
        summary["by_hour"] = dict(sorted(hours.items()))
        summary["chart"] = "bar"

    return summary


def _csv_safe(value):
    """Neutralise CSV / spreadsheet formula injection.

    A cell beginning with = + - @ (or a leading tab/CR) is interpreted as a
    formula by Excel/Sheets. Prefix such values with an apostrophe so they are
    rendered as literal text.
    """
    if value is None:
        return ""
    s = str(value)
    if s and s[0] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + s
    return s


def _cell_display(v):
    """Flatten an answer value for a single spreadsheet cell."""
    if isinstance(v, list):
        return "; ".join(str(x) for x in v)
    if v is None:
        return ""
    return v


def _apply_date_filter(query, qs_from, qs_to):
    """Apply ?from=YYYY-MM-DD and ?to=YYYY-MM-DD filters to a Response query."""
    if qs_from:
        try:
            dt = datetime.strptime(qs_from, "%Y-%m-%d")
            query = query.filter(Response.submitted_at >= dt)
        except ValueError:
            pass
    if qs_to:
        try:
            dt = datetime.strptime(qs_to, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(Response.submitted_at < dt)
        except ValueError:
            pass
    return query


# =========================================================================
@dashboard_bp.route("/forms/<int:form_id>/dashboard")
@login_required
def view_dashboard(form_id):
    form = require_form_ownership(form_id)
    return render_template("dashboard.html", form=form)


@dashboard_bp.route("/api/forms/<int:form_id>/analytics")
@login_required
def analytics(form_id):
    form = require_form_ownership(form_id)
    qs_from = request.args.get("from")
    qs_to = request.args.get("to")

    responses_q = form.responses
    responses_q = _apply_date_filter(responses_q, qs_from, qs_to)
    responses = responses_q.order_by(Response.submitted_at.asc()).all()

    total = len(responses)
    trend = Counter()
    for r in responses:
        day = r.submitted_at.strftime("%Y-%m-%d")
        trend[day] += 1
    trend_sorted = sorted(trend.items())

    # per-question summaries
    response_ids = [r.id for r in responses]
    question_summaries = []
    for q in form.questions:
        if q.type == "section":
            continue
        if response_ids:
            ans = Answer.query.filter(
                Answer.response_id.in_(response_ids),
                Answer.question_id == q.id,
            ).all()
        else:
            ans = []
        question_summaries.append(_summarize_question(q, ans))

    # completion rate: % of responses that answered all required questions
    req_qids = [q.id for q in form.questions if q.required and q.type != "section"]
    completed = 0
    for r in responses:
        ans_qids = {a.question_id for a in r.answers}
        if all(qid in ans_qids for qid in req_qids):
            completed += 1
    completion_rate = round(completed / total * 100, 1) if total else 0.0

    return jsonify({
        "form_id": form.id,
        "title": form.title,
        "total": total,
        "completion_rate": completion_rate,
        "trend": [{"date": d, "count": c} for d, c in trend_sorted],
        "questions": question_summaries,
    })


@dashboard_bp.route("/api/forms/<int:form_id>/responses")
@login_required
def list_responses(form_id):
    form = require_form_ownership(form_id)
    qs_from = request.args.get("from")
    qs_to = request.args.get("to")
    q = form.responses
    q = _apply_date_filter(q, qs_from, qs_to)
    rows = q.order_by(Response.submitted_at.desc()).all()
    return jsonify({
        "responses": [
            {
                "id": r.id,
                "submitted_at": r.submitted_at.isoformat(),
                "answers": {a.question_id: a.parsed_value() for a in r.answers},
            }
            for r in rows
        ]
    })


@dashboard_bp.route("/api/forms/<int:form_id>/responses/<int:rid>")
@login_required
def get_response(form_id, rid):
    form = require_form_ownership(form_id)
    r = Response.query.filter_by(id=rid, form_id=form.id).first_or_404()
    return jsonify({
        "id": r.id,
        "submitted_at": r.submitted_at.isoformat(),
        "answers": {a.question_id: a.parsed_value() for a in r.answers},
    })


@dashboard_bp.route("/api/forms/<int:form_id>/export.csv")
@login_required
def export_csv(form_id):
    form = require_form_ownership(form_id)
    questions = [q for q in form.questions if q.type != "section"]
    responses = form.responses.order_by(Response.submitted_at.asc()).all()

    out = io.StringIO()
    out.write("﻿")  # BOM for Excel UTF-8
    writer = csv.writer(out)
    header = ["#", "submitted_at"] + [q.text for q in questions]
    writer.writerow(header)

    for idx, r in enumerate(responses, 1):
        ans_map = {a.question_id: a.parsed_value() for a in r.answers}
        row = [idx, r.submitted_at.strftime("%Y-%m-%d %H:%M:%S")]
        for q in questions:
            row.append(_csv_safe(_cell_display(ans_map.get(q.id))))
        writer.writerow(row)

    resp = FlaskResponse(out.getvalue(), mimetype="text/csv; charset=utf-8")
    fname = f"form_{form.id}_responses.csv"
    resp.headers["Content-Disposition"] = f"attachment; filename=\"{fname}\""
    return resp


@dashboard_bp.route("/api/forms/<int:form_id>/export.xlsx")
@login_required
def export_xlsx(form_id):
    form = require_form_ownership(form_id)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "xlsx_unavailable",
                        "detail": "openpyxl ไม่ได้ติดตั้ง"}), 501

    questions = [q for q in form.questions if q.type != "section"]
    responses = form.responses.order_by(Response.submitted_at.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Responses"

    header = ["#", "submitted_at"] + [q.text for q in questions]
    ws.append(header)

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="4F46E5")
    for col_idx, _ in enumerate(header, 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = head_font
        c.fill = head_fill
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for idx, r in enumerate(responses, 1):
        ans_map = {a.question_id: a.parsed_value() for a in r.answers}
        row = [idx, r.submitted_at.strftime("%Y-%m-%d %H:%M:%S")]
        for q in questions:
            row.append(_csv_safe(_cell_display(ans_map.get(q.id))))
        ws.append(row)

    for col_idx, title in enumerate(header, 1):
        width = min(max(len(str(title)) + 2, 12), 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = FlaskResponse(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    fname = f"form_{form.id}_responses.xlsx"
    resp.headers["Content-Disposition"] = f"attachment; filename=\"{fname}\""
    return resp
